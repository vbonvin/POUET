"""
Useful functions and definitions
"""
import astropy.coordinates.angles as angles
#from astropy.table import Table
from astropy.time import Time
from astropy import units as u
import re
import getopt, sys, os
try:
	import openpyxl  # see http://openpyxl.readthedocs.org/en/latest/index.html
	noexcelimport = False
except:
	noexcelimport = True
from bisect import bisect_left
import pickle as pickle
import ephem
from configparser import SafeConfigParser
import importlib
import sys
import gzip
#import csv
import numpy as np

import obs

import logging
logger = logging.getLogger(__name__)



def takeclosest(dico, key, value):
	#todo: rename dico in dict
	"""
	.. warning:: I assume that dict[key] is sorted.

	Returns the dict value which dict[key] is closest to value.
	If two dict[key] are equally close to value, return the highest (i.e. latest).

	:param dico: python dictionary you want to sort
	:param key: dictionary key used for sorting
	:param value: target value,

	:return: index of the element in dico that is the closest to the target value

	.. note:: This is much faster than a simple min loop, although a bit more tedious to use.
	"""
	mylist = [elt[key] for elt in dico]

	pos = bisect_left(mylist, value)
	if pos == 0:
		return dico[0]
	if pos == len(mylist):
		return dico[-1]
	before = dico[pos - 1]
	after = dico[pos]
	if after[key] - value <= value - before[key]:
		return after
	else:
		return before


def hilite(string, status, bold):
	"""
	Helper to add colors and bold in the terminal

	:param string: string you want to color or bold
	:param status: boolean, if True then the text is colored in green, otherwise in red.
	:param bold: boolean, if True then the text is bolded
	:return:
	"""

	if not sys.stdout.isatty() : return '*'+string+'*'

	attr = []
	if status:
		# green
		attr.append('32')
	else:
		# red
		attr.append('31')
	if bold:
		attr.append('1')
	return '\x1b[%sm%s\x1b[0m' % (';'.join(attr), string)




def writepickle(obj, filepath, protocol=-1):
	"""
	I write your python object obj into a pickle file at filepath.
	If filepath ends with .gz, I'll use gzip to compress the pickle.

	:param obj: python container you want to compress
	:param filepath: string, path where the pickle will be written
	:param protocol: Leave protocol = -1 : I'll use the latest binary protocol of pickle.

	"""
	if os.path.splitext(filepath)[1] == ".gz":
		pkl_file = gzip.open(filepath, 'wb')
	else:
		pkl_file = open(filepath, 'wb')

	pickle.dump(obj, pkl_file, protocol)
	pkl_file.close()
	logger.debug("Wrote %s" % filepath)


def readpickle(filepath):
	"""
	I read a pickle file and return whatever object it contains.
	If the filepath ends with .gz, I'll unzip the pickle file.

	:param filepath: string, path of the pickle to load
	:return: object contained in the pickle
	"""
	if os.path.splitext(filepath)[1] == ".gz":
		pkl_file = gzip.open(filepath,'rb')
	else:
		pkl_file = open(filepath, 'rb')
	obj = pickle.load(pkl_file)
	pkl_file.close()
	logger.debug("Read %s" % filepath)
	return obj

def readconfig(configpath):
	"""
	Reads in a config file

	:param configpath: path of the configfile

	:return: configuration dictionary
	"""
	#todo: change SafeConfigParser to ConfigParser. Make sure logic is respected
	config = SafeConfigParser(allow_no_value=True)
	
	if not os.path.exists(configpath):
		raise RuntimeError("Config file '{}' does not exist!".format(configpath))
	logger.info("Reading config from '{}'...".format(configpath))
	config.read(configpath)
	
	return config

def grid_points(res_x=400,res_y=200):

	"""
	Maps the whole sky in right ascension and declination

	:param res_x: integer, number of points in right ascension
	:param res_y: integer, number of points in declination

	.. note:: the points are equally spaced.

	:return: numpy tuples containing right ascension points and declination points
	"""
	
	ra_i = 0.
	ra_f = 2*np.pi
	ra_step=(ra_f-ra_i)/res_x
	dec_i = -np.pi/2.
	dec_f = np.pi/2.
	dec_step=(dec_f-dec_i)/res_y
			
	ras = np.arange(ra_i+ra_step/2, ra_f, ra_step)
	decs= np.arange(dec_i+dec_step/2, dec_f, dec_step)
	return ras,decs

def elev2airmass(el, alt, threshold=10.):
	"""
	Converts the elevation to airmass.

	:param el: float, elevation in radians
	:param alt: float, altitude of the observer in meters
	:param threshold: maximum allowed airmass, will be returned if actual airmass exceeds the threshold

	:return: airmass

	.. note:: This is the code used for the Euler EDP at La Silla."""

	altitudeFactor = 0.00087 + alt*(-8.6664803e-8) # altitude factor

	cosz = np.cos(np.pi/2.-el)

	if(cosz< 0.1): # we do not compute Airmass for small value of cosz
		airmass = threshold
	else:
		airmass = (1.0 + altitudeFactor - altitudeFactor / (cosz * cosz)) / cosz

	return airmass

def check_value(var, flag):
	"""
	Check that a value is NaN, replace it with a given flag if True

	:param var: value to check against NaN
	:param flag: replacement value
	:return: processed value
	"""
	if np.isnan(var):
		var = flag 
	
	return var

def load_station(name):
	"""
	Load the parameters corresponding to an observation station

	:param name: string, name of the station. The corresponding file must be located in :py:`config`

	:return: station parameters
	"""
	module_name = "config.{}".format(name)
	station = importlib.import_module(module_name, package=None)
	return station

def time2hhmm(obstime):
	"""
	Concatenate a string HH MM SS or HH.MM.SS into an HHMM string

	:param obstime: string, HH MM SS or HH.MM.SS
	:return: HHMM string
	"""
	return (str(obstime).split(" ")[1]).split(".")[0][:-3]


def excelimport(filename, obsprogram=None):
	"""
	Wrapper around openpyxl to transform excel sheets into POUET.

	.. info:: This feature is still experimental and completely spreadsheet dependent. Requirex openpyxl installed, otherwise exits.

	Import an excel catalog into a list of observables

	:param filename: string, path to the excel file
	:param obsprogram: string, observing program associated to the catalog you are importing.

	:return: list of observables

	.. warning:: I directly read the excel values, I do NOT evaluate the formulas in them. It is up to the user to put the right mjd in the excel sheets.

	"""

	if noexcelimport:
		raise NotImplemented("Excel files cannot be imported at the moment - you need to install openpyxl")
	else:

		observables = []

		#### For BEBOP
		if obsprogram == 'bebop':
			"""
			special properties:

			phases : a list of dictionnaries : [{mjd, phase, hourafterstart }]
			comment : a string of comments (exptime, requested phase,...)
			internalobs : a boolean (0 or 1), allowing or not observability
			"""

			try:
				wb = openpyxl.load_workbook(filename,
				                            data_only=True)  # Read the excel spreadsheet, loading the values directly
				# ws = wb.active  # choose active sheet
				ws = wb['Sheet1']
			except:
				raise RuntimeError("Either %s does not exists, or it is not in .xlsx format !!" % filename)

			# Get tabler limits
			rows = ws.rows
			columns = ws.columns
			breakcolind = None
			breakrowind = None
			for ind, cell in enumerate(rows[1]):
				if cell.value == None:
					# breakcolind = cell.column
					breakcolind = rows[1][ind - 1].column
					break
				else:
					pass

			for ind, cell in enumerate(columns[0]):
				if cell.value == None:
					# breakrowind = cell.row
					breakrowind = columns[0][ind - 1].row
					break
				else:
					pass

			# Read only the non "None" data and put it in a table of dict, because fuck excel and fuck openpyxl.
			data = ws['A1':'%s%s' % (breakcolind, breakrowind)]
			"""
			Structure of the spreadsheet:
			Infos are from A1 to W2
			Datas are from A3 ro W30
			B1 : actual modified julian date
			A : name
			B : target
			C : comment
			I : observability
			J : requested phase
			M1 to W1 : mjd over the night
			M2 to W2 : corresponding time after night start, in hours
			M to W : phases
			"""

			phasesnames = ['M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W']
			values = {}
			for indr, row in enumerate(data):
				for indc, cell in enumerate(row):
					try:
						values[cell.address] = cell.value  # Apparently, this is an old version
					except:
						values[cell.coordinate] = cell.value

			for i in np.arange(3, 31):
				# create an observable object with the common properties
				name = values['A%s' % str(i)]

				coordinates = values['B%s' % str(i)]
				alpha = coordinates[0:2] + ':' + coordinates[2:4] + ':' + coordinates[4:6]
				delta = coordinates[7:9] + ':' + coordinates[9:11] + ':' + coordinates[11:13]
				if coordinates[6] == "S":
					delta = '-' + delta

				# add properties specific to this program
				## Tricky stuff here : the jdb in the excel sheet is the mjd + 0.5.
				phases = [{'mjd': values['%c%i' % (col, 1)] - 0.5, 'hourafterstart': values['%c%i' % (col, 2)],
				           'phase': values['%c%i' % (col, i)]} for col in phasesnames]
				attributes = {'phases': phases}
				# observable.phases = phases
				if values['I%s' % str(i)] == 'yes':
					attributes['internalobs'] = 1
				else:
					attributes['internalobs'] = 0

				observable = obs.Observable(name=name, obsprogram=obsprogram, alpha=alpha, delta=delta,
				                            attributes=attributes)

				comment = ''
				if values['C%s' % str(i)] is not None:
					comment = comment + values['C%s' % str(i)]
				if values['J%s' % str(i)] != '/':
					comment = comment + '\n' + 'Requested phase: ' + values['J%s' % str(i)]
				if comment != '':
					observable.comment = comment

				observables.append(observable)

		# TODO: check that the modified julian date corresponds to the ongoing night
		# TODO: assert that the above structure is correct ! (use keywords in the Info fields...?)

		if obsprogram == "transit":
			pass

		if obsprogram == "superwasp":
			# http://openpyxl.readthedocs.org/en/latest/optimized.html --- that will be useful for Amaury's monstruous spreadsheet
			"""
			special properties:

			phases : a list of dictionnaries : [{mjd, phase, hourafterstart }]
			comment : a string of comments (exptime, requested phase,...)
			internalobs : a boolean (0 or 1), allowing or not observability
			"""

			logger.info('reading %s...' % filename)
			try:
				wb = openpyxl.load_workbook(filename,
				                            data_only=True)  # Read the excel spreadsheet, loading the values directly
				ws = wb['Observations']  # choose active sheet
			except:
				raise RuntimeError("Either %s does not exists, or it is not in .xlsx format !!" % filename)

			logger.info('get tabler limits...')
			# Get tabler limits
			rows = ws.rows
			columns = ws.columns
			breakcolind = None
			breakrowind = None
			for ind, cell in enumerate(rows[1]):
				if cell.value == None:
					# breakcolind = cell.column
					breakcolind = rows[1][ind - 1].column
					break
				else:
					pass

			for ind, cell in enumerate(columns[0]):
				if cell.value == None:
					# breakrowind = cell.row
					breakrowind = columns[0][ind - 1].row
					break
				else:
					pass
			logger.info(breakrowind, breakcolind)
			sys.exit()

			# Read only the non "None" data and put it in a table of dict, because fuck excel and fuck openpyxl.
			data = ws['A1':'%s%s' % (breakcolind, breakrowind)]
			"""
			Structure of the spreadsheet:
			Infos are from A1 to W2
			Datas are from A3 ro W30
			B1 : actual modified julian date
			A : name
			B : target
			C : comment
			I : observability
			J : requested phase
			M1 to W1 : mjd over the night
			M2 to W2 : corresponding time after night start, in hours
			M to W : phases
			"""

		if obsprogram == "followup":
			pass

		return observables