"""
Useful functions and definitions
"""
import astropy.coordinates.angles as angles
#from astropy.table import Table
from astropy.time import Time
from astropy import units as u
import urllib2
import re
import getopt, sys, os
try:
	import openpyxl  # see http://openpyxl.readthedocs.org/en/latest/index.html
	noexcelimport = False
except:
	noexcelimport = True
import obs
from bisect import bisect_left
import cPickle as pickle
import ephem
from ConfigParser import SafeConfigParser

#import csv
import numpy as np

import logging
logger = logging.getLogger(__name__)


def reformat(coordinate, format):
	"""
	Transform a coordinate (hour, degree) in the format of your choice

	HHhDDdSSs <---> HH:DD:SS

	update : apparently useless, can use angle objects instead
	"""

	if 'm' in coordinate:
		if 'd' in coordinate:
			hd = coordinate.split('d')[0]
			m = coordinate.split('d')[1].split('m')[0]
		if 'h' in coordinate:
			hd = coordinate.split('h')[0]
			m = coordinate.split('h')[1].split('m')[0]
		s  = coordinate.split('m')[1].split('s')[0]

	elif ':' in coordinate:
		[hd, m, s] = coordinate.split(':')

	else:
		raise ValueError("%s, Unknown coordinate input format!" %coordinate)

	if format == 'numeric':
		return "%s:%s:%s" % (hd, m, s)

	elif format == 'alphabetic_degree':
		return "%sd%sm%ss" % (hd, m, s)

	elif format == 'alphabetic_hour':
		return "%sh%sm%ss" % (hd, m, s)

	else:
		raise ValueError("%s, Unknown coordinate output format!" %format)


def takeclosest(dico, key, value):
	"""
	Assumes dict[key] is sorted. Returns the dict value which dict[key] is closest to value.
	If two dict[key] are equally close to value, return the highest (i.e. latest).
	This is much faster than a simple min loop, although a bit more tedious to use.
	"""
	mylist = [elt[key] for elt in dico]

	pos = bisect_left(mylist, value)
	if pos == 0:
		return dico[0]
	if pos == len(mylist):
		return dico[-1]
	before = dico[pos - 1]
	after = dico[pos]
	if after[key] - value <= value - before[key]:
		return after
	else:
		return before


def hilite(string, status, bold):
	'''Graphism: colors and bold in the terminal'''

	if not sys.stdout.isatty() : return '*'+string+'*'

	attr = []
	if status:
		# green
		attr.append('32')
	else:
		# red
		attr.append('31')
	if bold:
		attr.append('1')
	return '\x1b[%sm%s\x1b[0m' % (';'.join(attr), string)

def excelimport(filename, obsprogram=None):
	if noexcelimport:
		raise NotImplemented("Excel files cannot be imported at the moment")
	else:

		"""
		Import an excel catalog(...) into a list of observables
		I directly read the excel values, I do NOT evaluate the formulas in them.
		It is up to you to put the right mjd in the excel sheets.
	
		Warning : NEVER use the coordinates from an excel sheet to create an rdb night planning. ALWAYS use the rdb catalogs loaded in the edp. And double check the distance to moon !
		"""

		observables = []

		#### For BEBOP
		if obsprogram == 'bebop':
			"""
			special properties:
	
			phases : a list of dictionnaries : [{mjd, phase, hourafterstart }]
			comment : a string of comments (exptime, requested phase,...)
			internalobs : a boolean (0 or 1), allowing or not observability
			"""

			try:
				wb = openpyxl.load_workbook(filename, data_only=True)  # Read the excel spreadsheet, loading the values directly
				#ws = wb.active  # choose active sheet
				ws = wb['Sheet1']
			except:
				raise RuntimeError("Either %s does not exists, or it is not in .xlsx format !!" % filename)

			# Get tabler limits
			rows = ws.rows
			columns = ws.columns
			breakcolind = None
			breakrowind = None
			for ind, cell in enumerate(rows[1]):
				if cell.value == None:
					#breakcolind = cell.column
					breakcolind = rows[1][ind-1].column
					break
				else:
					pass

			for ind, cell in enumerate(columns[0]):
				if cell.value == None:
					#breakrowind = cell.row
					breakrowind = columns[0][ind-1].row
					break
				else:
					pass

			# Read only the non "None" data and put it in a table of dict, because fuck excel and fuck openpyxl.
			data = ws['A1':'%s%s' % (breakcolind, breakrowind)]
			"""
			Structure of the spreadsheet:
			Infos are from A1 to W2
			Datas are from A3 ro W30
			B1 : actual modified julian date
			A : name
			B : target
			C : comment
			I : observability
			J : requested phase
			M1 to W1 : mjd over the night
			M2 to W2 : corresponding time after night start, in hours
			M to W : phases
			"""

			phasesnames = ['M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W']
			values = {}
			for indr, row in enumerate(data):
				for indc, cell in enumerate(row):
					try:
						values[cell.address] = cell.value # Apparently, this is an old version
					except:
						values[cell.coordinate] = cell.value

			for i in np.arange(3, 31):
				# create an observable object with the common properties
				name = values['A%s' % str(i)]

				coordinates = values['B%s' % str(i)]
				alpha = coordinates[0:2]+':'+coordinates[2:4]+':'+coordinates[4:6]
				delta = coordinates[7:9]+':'+coordinates[9:11]+':'+coordinates[11:13]
				if coordinates[6] == "S":
					delta = '-'+delta

				# add properties specific to this program
				## Tricky stuff here : the jdb in the excel sheet is the mjd + 0.5.
				phases = [{'mjd': values['%c%i' % (col, 1)]-0.5, 'hourafterstart': values['%c%i' % (col, 2)], 'phase': values['%c%i' % (col, i)]} for col in phasesnames]
				attributes = {'phases': phases}
				#observable.phases = phases
				if values['I%s' % str(i)] == 'yes':
					attributes['internalobs'] = 1
				else:
					attributes['internalobs'] = 0

				observable = obs.Observable(name=name, obsprogram=obsprogram, alpha=alpha, delta=delta,
					attributes=attributes)

				comment = ''
				if values['C%s' % str(i)] is not None:
					comment = comment + values['C%s' % str(i)]
				if values['J%s' % str(i)] != '/':
					comment = comment + '\n' + 'Requested phase: ' + values['J%s' % str(i)]
				if comment != '':
					observable.comment = comment

				observables.append(observable)

			#TODO: check that the modified julian date corresponds to the ongoing night
			#TODO: assert that the above structure is correct ! (use keywords in the Info fields...?)

		if obsprogram == "transit":
			pass

		if obsprogram == "superwasp":
			# http://openpyxl.readthedocs.org/en/latest/optimized.html --- that will be useful for Amaury's monstruous spreadsheet
			"""
			special properties:
	
			phases : a list of dictionnaries : [{mjd, phase, hourafterstart }]
			comment : a string of comments (exptime, requested phase,...)
			internalobs : a boolean (0 or 1), allowing or not observability
			"""

			logger.info('reading %s...' % filename)
			try:
				wb = openpyxl.load_workbook(filename, data_only=True)  # Read the excel spreadsheet, loading the values directly
				ws = wb['Observations']  # choose active sheet
			except:
				raise RuntimeError("Either %s does not exists, or it is not in .xlsx format !!" % filename)

			logger.info('get tabler limits...')
			# Get tabler limits
			rows = ws.rows
			columns = ws.columns
			breakcolind = None
			breakrowind = None
			for ind, cell in enumerate(rows[1]):
				if cell.value == None:
					#breakcolind = cell.column
					breakcolind = rows[1][ind-1].column
					break
				else:
					pass

			for ind, cell in enumerate(columns[0]):
				if cell.value == None:
					#breakrowind = cell.row
					breakrowind = columns[0][ind-1].row
					break
				else:
					pass
			logger.info(breakrowind, breakcolind)
			sys.exit()


			# Read only the non "None" data and put it in a table of dict, because fuck excel and fuck openpyxl.
			data = ws['A1':'%s%s' % (breakcolind, breakrowind)]
			"""
			Structure of the spreadsheet:
			Infos are from A1 to W2
			Datas are from A3 ro W30
			B1 : actual modified julian date
			A : name
			B : target
			C : comment
			I : observability
			J : requested phase
			M1 to W1 : mjd over the night
			M2 to W2 : corresponding time after night start, in hours
			M to W : phases
			"""

		if obsprogram == "followup":
			pass

		return observables

def rdbimport(filepath, namecol=1, alphacol=2, deltacol=3, startline=1, obsprogram="None", verbose=False):
	"""
	Import an rdb catalog into a list of observables
	THIS SHOULD BE IN UTIL !!
	"""

	logger.info("Reading \"%s\"..." % (os.path.basename(filepath)))
	rdbfile = open(filepath, "r")
	rdbfilelines = rdbfile.readlines()[startline:] # we directly "skip" the first lines of eventual headers
	rdbfile.close()

	observables = []

	for ind, line in enumerate(rdbfilelines) :

		if line[0] == "-" or line[0] == "#":
			continue

		if len(line.strip()) < 5:
			logger.debug("Skipping empty line %i : %s" % (ind+startline, repr(line)))
			continue

		elements = line.split()

		name = str(elements[namecol-1])
		alpha = str(elements[alphacol-1])
		delta = str(elements[deltacol-1])

		# This is the minimal stuff necessary to define an observable. Now, let's go into the per-obsprogram details

		if obsprogram not in ["lens", "transit", "bebop", "superwasp", "followup", "703", "714"]:
			observables.append(Observable(name=name, obsprogram=obsprogram, alpha=alpha, delta=delta))

		if obsprogram == "lens":
			minangletomoon = 30
			maxairmass = 1.5
			exptime = 35*60 #approx 35 minutes per lens
			observables.append(Observable(name=name, obsprogram=obsprogram, alpha=alpha, delta=delta, minangletomoon=minangletomoon, maxairmass=maxairmass, exptime=exptime))

		if obsprogram == "transit":
			pass

		if obsprogram == "bebop":
			pass

		if obsprogram == "superwasp":
			pass

		if obsprogram == "followup":
			pass

		if obsprogram == "703":
			pass

		if obsprogram == "714":
			pass

	return observables



def writepickle(obj, filepath, verbose=True, protocol = -1):
	"""
	I write your python object obj into a pickle file at filepath.
	If filepath ends with .gz, I'll use gzip to compress the pickle.
	Leave protocol = -1 : I'll use the latest binary protocol of pickle.
	"""
	if os.path.splitext(filepath)[1] == ".gz":
		pkl_file = gzip.open(filepath, 'wb')
	else:
		pkl_file = open(filepath, 'wb')

	pickle.dump(obj, pkl_file, protocol)
	pkl_file.close()
	logger.debug("Wrote %s" % filepath)


def readpickle(filepath, verbose=True):
	"""
	I read a pickle file and return whatever object it contains.
	If the filepath ends with .gz, I'll unzip the pickle file.
	"""
	if os.path.splitext(filepath)[1] == ".gz":
		pkl_file = gzip.open(filepath,'rb')
	else:
		pkl_file = open(filepath, 'rb')
	obj = pickle.load(pkl_file)
	pkl_file.close()
	logger.debug("Read %s" % filepath)
	return obj

def readconfig(configpath):
	"""
	Reads in a config file
	"""
	config = SafeConfigParser(allow_no_value=True)
	
	if not os.path.exists(configpath):
		raise RuntimeError("Config file '{}' does not exist!".format(configpath))
	logger.info("Reading config from '{}'...".format(configpath))
	config.read(configpath)
	
	return config
